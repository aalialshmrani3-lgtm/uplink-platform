"""
Excel Export Service for Strategic Analysis Reports
Generates comprehensive Excel workbooks from strategic analysis data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Any, List

class StrategicAnalysisExcelExporterEN:
    def __init__(self):
        self.workbook = Workbook()
        self.workbook.remove(self.workbook.active)  # Remove default sheet
        
        # Define styles
        self.header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=12)
        self.title_font = Font(bold=True, size=14, color="1E40AF")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def generate_report(self, analysis_data: Dict[str, Any], output_path: str) -> str:
        """
        Generate Excel report from strategic analysis data
        
        Args:
            analysis_data: Dictionary containing strategic analysis results
            output_path: Path where Excel file should be saved
            
        Returns:
            Path to generated Excel file
        """
        # Create sheets
        self._create_summary_sheet(analysis_data)
        self._create_ici_sheet(analysis_data)
        self._create_insights_sheet(analysis_data)
        self._create_roadmap_sheet(analysis_data)
        self._create_investment_sheet(analysis_data)
        self._create_critical_path_sheet(analysis_data)
        
        # Save workbook
        self.workbook.save(output_path)
        return output_path
    
    def _create_summary_sheet(self, data: Dict[str, Any]):
        """Create Executive Summary sheet"""
        ws = self.workbook.create_sheet("Executive Summary")
        
        # Title
        ws['A1'] = "UPLINK 5.0 - Strategic Analysis Report"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Project Title
        ws['A3'] = "Project:"
        ws['A3'].font = Font(bold=True)
        ws['B3'] = data.get('project_title', 'Untitled Project')
        
        # Metadata
        row = 5
        metadata = [
            ("Report Date:", datetime.now().strftime('%Y-%m-%d %H:%M')),
            ("ICI Score:", f"{data.get('ici_score', 0)}/100"),
            ("IRL Score:", f"{data.get('irl_score', 0)}/100"),
            ("Success Probability:", f"{data.get('success_probability', 0)*100:.1f}%"),
            ("Risk Level:", data.get('risk_level', 'N/A')),
            ("Investor Appeal:", data.get('investor_appeal', 'N/A'))
        ]
        
        for label, value in metadata:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Executive Summary
        row += 2
        ws[f'A{row}'] = "Executive Summary:"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        ws[f'A{row}'] = data.get('executive_summary', 'No summary available.')
        ws.merge_cells(f'A{row}:D{row+3}')
        ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
        
        # Key Dimensions Table
        row += 5
        ws[f'A{row}'] = "Key Dimensions"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        dimensions = data.get('dimensions', {})
        dim_data = [
            ("Dimension", "Score"),
            ("Success Probability", f"{dimensions.get('success_probability', 0):.1f}"),
            ("Market Fit", f"{dimensions.get('market_fit', 0):.1f}"),
            ("Execution Readiness", f"{dimensions.get('execution_readiness', 0):.1f}"),
            ("Investor Readiness", f"{dimensions.get('investor_readiness', 0):.1f}"),
            ("Financial Sustainability", f"{dimensions.get('financial_sustainability', 0):.1f}")
        ]
        
        for col_idx, (label, value) in enumerate(dim_data, start=1):
            if col_idx == 1:  # Header row
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].fill = self.header_fill
                ws[f'B{row}'].fill = self.header_fill
                ws[f'A{row}'].font = self.header_font
                ws[f'B{row}'].font = self.header_font
            else:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                ws[f'A{row}'].border = self.border
                ws[f'B{row}'].border = self.border
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
    
    def _create_ici_sheet(self, data: Dict[str, Any]):
        """Create Innovation Confidence Index sheet"""
        ws = self.workbook.create_sheet("ICI Analysis")
        
        # Title
        ws['A1'] = "Innovation Confidence Index (ICI)"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:C1')
        
        # ICI Score
        ws['A3'] = "ICI Score:"
        ws['A3'].font = Font(bold=True)
        ws['B3'] = f"{data.get('ici_score', 0)}/100"
        ws['B3'].font = Font(size=14, bold=True, color="1E40AF")
        
        ws['A4'] = "Level:"
        ws['A4'].font = Font(bold=True)
        ws['B4'] = data.get('ici_level', 'N/A')
        
        # Dimensions breakdown
        row = 6
        ws[f'A{row}'] = "Dimension Breakdown"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        dimensions = data.get('dimensions', {})
        dim_list = [
            ("Dimension", "Score", "Weight"),
            ("Success Probability", dimensions.get('success_probability', 0), "20%"),
            ("Market Fit", dimensions.get('market_fit', 0), "20%"),
            ("Execution Readiness", dimensions.get('execution_readiness', 0), "20%"),
            ("Investor Readiness", dimensions.get('investor_readiness', 0), "20%"),
            ("Financial Sustainability", dimensions.get('financial_sustainability', 0), "20%")
        ]
        
        for idx, (dim, score, weight) in enumerate(dim_list):
            if idx == 0:  # Header
                ws[f'A{row}'] = dim
                ws[f'B{row}'] = score
                ws[f'C{row}'] = weight
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].fill = self.header_fill
                    ws[f'{col}{row}'].font = self.header_font
            else:
                ws[f'A{row}'] = dim
                ws[f'B{row}'] = score if isinstance(score, str) else f"{score:.1f}"
                ws[f'C{row}'] = weight
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{row}'].border = self.border
            row += 1
        
        # Create bar chart
        chart = BarChart()
        chart.title = "ICI Dimensions"
        chart.y_axis.title = "Score"
        chart.x_axis.title = "Dimension"
        
        data_ref = Reference(ws, min_col=2, min_row=8, max_row=12)
        cats_ref = Reference(ws, min_col=1, min_row=9, max_row=12)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(cats_ref)
        
        ws.add_chart(chart, "E6")
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_insights_sheet(self, data: Dict[str, Any]):
        """Create CEO Insights sheet"""
        ws = self.workbook.create_sheet("CEO Insights")
        
        # Title
        ws['A1'] = "CEO Insights"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Header row
        row = 3
        headers = ["#", "Insight", "Impact", "Priority"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        # Insights data
        insights = data.get('ceo_insights', [])
        row += 1
        
        for idx, insight in enumerate(insights[:20], start=1):  # Limit to top 20
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=insight.get('insight', ''))
            ws.cell(row=row, column=3, value=insight.get('impact', ''))
            ws.cell(row=row, column=4, value=insight.get('priority', 'MEDIUM'))
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 15
    
    def _create_roadmap_sheet(self, data: Dict[str, Any]):
        """Create Actionable Roadmap sheet"""
        ws = self.workbook.create_sheet("Actionable Roadmap")
        
        # Title
        ws['A1'] = "Actionable Roadmap"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:E1')
        
        # Header row
        row = 3
        headers = ["Step", "Title", "Description", "Timeline", "Priority"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        # Roadmap data
        roadmap = data.get('roadmap', {})
        steps = roadmap.get('steps', [])
        row += 1
        
        for idx, step in enumerate(steps, start=1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=step.get('title', ''))
            ws.cell(row=row, column=3, value=step.get('description', ''))
            ws.cell(row=row, column=4, value=step.get('timeline', 'N/A'))
            ws.cell(row=row, column=5, value=step.get('priority', 'MEDIUM'))
            
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
    
    def _create_investment_sheet(self, data: Dict[str, Any]):
        """Create Investment Analysis sheet"""
        ws = self.workbook.create_sheet("Investment Analysis")
        
        # Title
        ws['A1'] = "Investment Analysis"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Investment Summary
        investment = data.get('investment', {})
        row = 3
        
        summary_data = [
            ("Valuation Range:", investment.get('valuation_range', 'N/A')),
            ("Funding Potential:", investment.get('funding_potential', 'N/A')),
            ("IRL Score:", f"{data.get('irl_score', 0)}/100")
        ]
        
        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Recommended Investors
        row += 2
        ws[f'A{row}'] = "Recommended Investors"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers = ["Investor Type", "Probability", "Amount", "Timeline"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        row += 1
        recommended_investors = investment.get('recommended_investors', [])
        
        for investor in recommended_investors[:10]:  # Limit to top 10
            ws.cell(row=row, column=1, value=investor.get('name', 'Unknown'))
            ws.cell(row=row, column=2, value=f"{investor.get('probability', 0)*100:.0f}%")
            ws.cell(row=row, column=3, value=investor.get('amount', 'N/A'))
            ws.cell(row=row, column=4, value=investor.get('timeline', 'N/A'))
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
    
    def _create_critical_path_sheet(self, data: Dict[str, Any]):
        """Create Critical Path sheet"""
        ws = self.workbook.create_sheet("Critical Path")
        
        # Title
        ws['A1'] = "Critical Path to Success"
        ws['A1'].font = self.title_font
        ws.merge_cells('A1:D1')
        
        # Header row
        row = 3
        headers = ["Milestone", "Title", "Description", "Duration"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.border = self.border
        
        # Critical path data
        critical_path = data.get('critical_path', [])
        row += 1
        
        for idx, milestone in enumerate(critical_path, start=1):
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=2, value=milestone.get('title', ''))
            ws.cell(row=row, column=3, value=milestone.get('description', ''))
            ws.cell(row=row, column=4, value=milestone.get('duration', 'N/A'))
            
            for col in range(1, 5):
                ws.cell(row=row, column=col).border = self.border
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical='top')
            
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 10
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 15


def export_to_excel_en(analysis_data: Dict[str, Any], output_path: str) -> str:
    """
    Main function to export strategic analysis to English Excel
    
    Args:
        analysis_data: Strategic analysis data dictionary
        output_path: Path where Excel file should be saved
        
    Returns:
        Path to generated Excel file
    """
    exporter = StrategicAnalysisExcelExporterEN()
    return exporter.generate_report(analysis_data, output_path)


if __name__ == "__main__":
    # Test with sample data
    sample_data = {
        'project_title': 'Test Innovation Project',
        'ici_score': 65.5,
        'irl_score': 58.2,
        'success_probability': 0.72,
        'risk_level': 'MEDIUM',
        'investor_appeal': 'HIGH',
        'ici_level': 'Moderate',
        'executive_summary': 'This project shows promising potential with strong market fit and execution readiness.',
        'dimensions': {
            'success_probability': 72.0,
            'market_fit': 68.5,
            'execution_readiness': 65.0,
            'investor_readiness': 58.2,
            'financial_sustainability': 52.0
        },
        'ceo_insights': [
            {
                'insight': 'Strong market demand validates product-market fit',
                'impact': 'High positive impact on success probability',
                'priority': 'HIGH'
            }
        ],
        'roadmap': {
            'steps': [
                {
                    'title': 'Secure Seed Funding',
                    'description': 'Approach angel investors and accelerator programs',
                    'timeline': '2-3 months',
                    'priority': 'HIGH'
                }
            ]
        },
        'investment': {
            'valuation_range': '5M - 10M SAR',
            'funding_potential': '1M - 2M SAR',
            'recommended_investors': [
                {
                    'name': 'Angel Investors',
                    'probability': 0.35,
                    'amount': '500K SAR',
                    'timeline': '1 month'
                }
            ]
        },
        'critical_path': [
            {
                'title': 'Product Development',
                'description': 'Complete MVP and initial testing',
                'duration': '3-4 months'
            }
        ]
    }
    
    output = export_to_excel_en(sample_data, '/tmp/test_report_en.xlsx')
    print(f"✅ English Excel generated: {output}")
